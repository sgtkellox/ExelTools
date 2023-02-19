import pandas as pd
from openpyxl import load_workbook

pathTable1 = r"C:\Users\felix\Documents\fuatsTableTest.xlsx"
pathTable2 = r"C:\Users\felix\Documents\supp3short.xlsx"

pathOut = r"C:\Users\felix\Desktop\fuatTest\result.xlsx"

table1 = pd.read_excel(pathTable1)
table2 = pd.read_excel(pathTable2)

#wb = load_workbook(filename = pathTable2)

#print(wb.sheetnames)
#print(table2['SITE_+/-7_AA'].values)

final = pd.DataFrame(columns=table1.columns)
print(table1.columns)

#final.to_excel(pathOut, index = False)

for ind in table1.index:
    val1 = table1['peptide'][ind]
    #print("read val at" + str(ind)+ " "+ val1)
    concatVal = ""
    if "s" in val1: 
        val1Split = val1.split("s")
        valSplitLen = len(val1Split[1])
        if valSplitLen>6:
            subVal1 = val1Split[1][0:7]
            lenghtStringBeforeS = len(val1Split[0])
            subVal2 = val1Split[0][lenghtStringBeforeS-7:lenghtStringBeforeS]
            concatVal = subVal2+subVal1
            print(concatVal)
        else:
            subVal1 = val1Split[1][0:valSplitLen]
            lenghtStringBeforeS = len(val1Split[0])
            subVal2 = val1Split[0][lenghtStringBeforeS-valSplitLen:lenghtStringBeforeS]
            concatVal = subVal2+subVal1
    elif "t" in val1:
        val1Split = val1.split("t")
        valSplitLen = len(val1Split[1])
        if valSplitLen>6:
            subVal1 = val1Split[1][0:7]
            lenghtStringBeforeT = len(val1Split[0])
            subVal2 = val1Split[0][lenghtStringBeforeT-7:lenghtStringBeforeT]
            concatVal = subVal2+subVal1
        else:
            subVal1 = val1Split[1][0:valSplitLen]
            lenghtStringBeforeT = len(val1Split[0])
            subVal2 = val1Split[0][lenghtStringBeforeT-valSplitLen:lenghtStringBeforeT]
            concatVal = subVal2+subVal1
    if not concatVal == "" and concatVal in table2['SITE_+/-7_AA'].values:
        #final.loc[len(final)] = table1.iloc[[ind]]
        final = final.append(table1.iloc[[ind]],ignore_index=True)

final.to_excel(pathOut, index = False)


